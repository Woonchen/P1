from openpyxl import Workbook, load_workbook
from datetime import datetime

def saveToExcel(time, buy, sell, ratio, memo):

    # Excel 文件名
    filename = '/Users/huangwencheng/Desktop/P1/historydata.xlsx'

    # 要寫入的計入
    new_record = {'Time': time, 'Buy': buy, 'Sell': sell, 'Ratio': ratio, 'Memo': memo}

    # 建立或載入 Excel 工作簿
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()

    # 建立工作表
    ws = wb.active

    # 將新計入插入到工作表的第一行
    ws.insert_rows(1)
    for col, (key, value) in enumerate(new_record.items(), start=1):
        ws.cell(row=1, column=col, value=key)
        ws.cell(row=2, column=col, value=value)

    # 保存 Excel 
    wb.save(filename)

# 獲取當前時間
#current_time = datetime.now()
# 格式化當前時間
#formatted_time = current_time.strftime("%Y-%m-%d %H:%M:%S")


#print("當前時間:", formatted_time)
#saveToExcel(formatted_time, 2.333, 3.444, 4)
